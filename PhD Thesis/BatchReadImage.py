# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 11:28:11 2024

@author: SG
"""
import cv2
import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import time


os.chdir("C:/Users/Laboratorio/AutoMeasureLCwithPolorizationCamera/monitor/monitor_mor")
excelName=input("please input a filename for saving data:")
start_time = time.time()
filename=0
maxinumber=50 # this value is for the upper bound in the reading imgae loop
while filename<maxinumber+1:
    
    im=cv2.imread(str(filename)+".tif")
    im1=im[:,:,1] # green channel
    print(filename) 
    im45withlabel=im1[1150:1850,250:1050]
    im_45withlabel=im1[100:800,1470:2170]

    im45=im45withlabel[350:500,230:400]
    im_45=im_45withlabel[350:500,230:400] # adjust these arraies to make the light in the center
    plt.subplot(2, 2, 1)
    plt.imshow(im45withlabel)
    plt.title('+45')
    plt.subplot(2, 2, 2)
    plt.imshow(im_45withlabel)
    plt.title('-45')

    plt.subplot(2, 2, 3)
    plt.imshow(im45)
    plt.title('+45')
    plt.subplot(2, 2, 4)
    plt.imshow(im_45)
    plt.title('-45')


    def calIntensity(row, col,x,y,r,pol):
        count=0
        sumintensity=0 # for sum work here it should be 0
        for i in range(rows):  
            for j in range(cols):
                distance_squared = (i - x0) ** 2 + (j - y0) ** 2
                if distance_squared < r ** 2:
                    if pol=="45":
                        sumintensity += im45[i, j]                    
                    else:
                        sumintensity += im_45[i, j]
                    count+=1
                    return sumintensity/count
                        
        print(f"The sum of the intensity id {sumintensity} and the number of the pixel is {count} ")
        print(f"Then the average of the intensity is {sumintensity/count:.2f}")
        
    # print("For +45")
    rows, cols = np.shape(im45)
    x0,y0=rows/2,cols/2 
    r=20
    cal45=calIntensity(rows, cols, x0, y0, r,"45")
    # print("For -45")
    rows_45, cols_45 = np.shape(im_45)
    x0_45,y0_45=rows_45/2,cols_45/2 
    cal_45=calIntensity(rows_45, cols_45, x0_45, y0_45, r,"-45")

    
    if os.path.exists(excelName+ ".xlsx"):
        wb=load_workbook(excelName+r".xlsx")
        ws=wb.active
        ws['D1'] = "I (45)-BG"
        ws['E1'] = "I (-45)-BG"
        ws['F1'] = "Sum"
        ws['H1'] = "I_c"
        ws['I1'] = "I_p"
        ws['J1'] = "SQRT"
        ws['K1'] = "2*ATAN"
        ws['L1'] = "phase(π)"
        ws['M1'] = "unwrapped phase(π)"
        ws.cell(filename+2,2,cal45)
        ws.cell(filename+2,3,cal_45)
        i=2
        while i<maxinumber+3:
            ws[f'D{i}'] = f"=B{i}-MIN(B:B)"
            ws[f'E{i}'] = f"=C{i}-MIN(C:C)"
            ws[f'F{i}'] = f"=E{i}+D{i}"
            ws[f'H{i}'] = f"=D{i}/F{i}"
            ws[f'I{i}'] = f"=E{i}/F{i}"
            ws[f'J{i}'] = f"=SQRT(H{i}/I{i})"
            ws[f'K{i}'] = f"=2*ATAN(J{i})"
            ws[f'L{i}'] = f"=K{i}/PI()"
            
            i+=1
        wb.save(excelName+r".xlsx")
    else:
        wb=Workbook()
        ws=wb.active
        ws['D1'] = "I (45)-BG"
        ws['E1'] = "I (-45)-BG"
        ws['F1'] = "Sum"
        ws['H1'] = "I_c"
        ws['I1'] = "I_p"
        ws['J1'] = "SQRT"
        ws['K1'] = "2*ATAN"
        ws['L1'] = "phase(π)"
        ws['M1'] = "unwrapped phase(π)"
        ws.cell(1,1,"ImageName")
        ws.cell(1,2,"pol=+45")
        ws.cell(1,3,"pol=-45")
        ws.cell(1,7,"Voltage")
        vols=np.arange(0,5.1,0.1) # the number of how many voltage, and +2 is the number of row
        for k  in range(len(vols)):
              ws.cell(k+2, 1, vols[k]*20)
              ws.cell(k+2, 7, vols[k])
        ws.cell(filename+2,2,cal45)
        ws.cell(filename+2,3,cal_45)
        i=2
        while i<maxinumber+3:
            ws[f'D{i}'] = f"=B{i}-MIN(B:B)"
            ws[f'E{i}'] = f"=C{i}-MIN(C:C)"
            ws[f'F{i}'] = f"=E{i}+D{i}"
            ws[f'H{i}'] = f"=D{i}/F{i}"
            ws[f'I{i}'] = f"=E{i}/F{i}"
            ws[f'J{i}'] = f"=SQRT(H{i}/I{i})"
            ws[f'K{i}'] = f"=2*ATAN(J{i})"
            ws[f'L{i}'] = f"=K{i}/PI()"
            i+=1
        wb.save(excelName+r".xlsx")
    filename+=1

end_time = time.time()
execution_time = end_time - start_time
print("Execution time:", execution_time, "seconds")