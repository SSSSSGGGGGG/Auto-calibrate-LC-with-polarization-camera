# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 11:28:11 2024

@author: SG
"""

import os
from openpyxl import workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart import Reference as opyxlReference

os.chdir("C:/Users/Laboratorio/scitificCalculation/imageHub/g")
filename="g.xlsx"
wb=load_workbook(filename)
ws=wb.active
ws['D1'] = "I (c)-BG"
ws['E1'] = "I (p)-BG"
ws['F1'] = "Sum"
ws['H1'] = "I_c"
ws['I1'] = "I_p"
ws['J1'] = "SQRT"
ws['K1'] = "2*ATAN"
ws['L1'] = "phase(π)"

i=2
while i<103:
    ws[f'D{i}'] = f"=B{i}-MIN(B:B)"
    ws[f'E{i}'] = f"=C{i}-MIN(C:C)"
    ws[f'F{i}'] = f"=E{i}+D{i}"
    ws[f'H{i}'] = f"=D{i}/F{i}"
    ws[f'I{i}'] = f"=E{i}/F{i}"
    ws[f'J{i}'] = f"=SQRT(H{i}/I{i})"
    ws[f'K{i}'] = f"=2*ATAN(J{i})"
    ws[f'L{i}'] = f"=K{i}/PI()"
    i+=1
# ws.cell(2,4,"=")
chart=ScatterChart()
x=opyxlReference(ws,min_col=7,min_row=2,max_row=102)
y1=Reference(ws,min_col=8,min_row=2,max_row=102)
y2=Reference(ws,min_col=9,min_row=2,max_row=102)
series1=Series(y1,x)
series2=Series(y2,x)
chart.series.append(series1)
chart.series.append(series2)
ws.add_chart(chart, "N2")
wb.save(filename)